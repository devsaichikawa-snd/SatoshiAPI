from typing import Any
from openpyxl import load_workbook

from common.util import is_none_or_empty


def read_workbook(file_path: str):
    """指定したExcel Bookを読み込む"""
    if file_path is None:
        raise ValueError("ファイルパスが存在しません。")
    wb = load_workbook(file_path, data_only=True)
    return wb


def save_and_close_book(wb, file_path=None, save_flag: bool = False):
    """指定したExcel Bookを保存・閉じる"""
    if save_flag:
        if file_path:
            wb.save(file_path)
        else:
            raise ValueError("保存する場合はファイルパスを指定してください。")
    wb.close()
    print("workbookを閉じました。")


def read_sheet(wb, sheet_name=None):
    """指定したExcel Sheetを開く"""
    if sheet_name is None:
        ws = wb.active
    else:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            raise ValueError(
                f"指定されたシート名 '{sheet_name}' は存在しません。"
            )
    return ws


def read_cell(ws, row: int, column: int, for_database: bool = False):
    """cellの値を取得する"""
    if row < 1 or column < 1:
        raise ValueError("行と列は1以上である必要があります。")

    value = ws.cell(row=row, column=column).value

    # 取得したセル値がNoneまたは空文字であるかどうかをチェックする
    if is_none_or_empty(value):
        if for_database:
            # DB-Insertで使用するのでNoneを返す
            return None
        else:
            # データクレンジングなので、空文字で返す
            return ""

    return str(value)


def write_cell(ws, row: int, column: int, value: Any):
    """cellの値を書き込む"""
    if row < 1 or column < 1:
        raise ValueError("行と列は1以上である必要があります。")

    ws.cell(row=row, column=column).value = value


def excel_to_pdf():
    """ExcelファイルをPDFファイルに変換する"""
    pass
